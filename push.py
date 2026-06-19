# LOCAL: uploads videos + builds schedule queue
import os
import subprocess

import openpyxl
from openpyxl.utils import get_column_letter
import cloudinary
import cloudinary.uploader
from dotenv import load_dotenv

XLSX = "schedule.xlsx"
SKIP_STATUSES = {"posted", "published", "missed", "failed", "posting"}

# ---------------------------------------------------------------------------
# 1. CONFIG
# ---------------------------------------------------------------------------
load_dotenv("config.env")
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
)


# ---------------------------------------------------------------------------
# COLUMN MAPPING
# ---------------------------------------------------------------------------
def map_columns(ws):
    headers = {}  # column letter -> lowercased header text
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() != "":
            headers[cell.column_letter] = str(cell.value).strip().lower()

    def find(pred):
        for letter, text in headers.items():
            if pred(text):
                return letter
        return None

    return {
        "name": find(lambda h: h in ("reel name", "name")),
        "caption": find(lambda h: "caption" in h),
        "datetime": find(lambda h: "date" in h),
        "status": find(lambda h: h == "status"),
        "video_url": find(lambda h: h == "video url"),
    }


def run_git(args):
    """Run a git command in the project folder, print command + output.
    Returns the CompletedProcess."""
    print(f"$ {' '.join(args)}")
    proc = subprocess.run(args, capture_output=True, text=True)
    if proc.stdout:
        print(proc.stdout.rstrip())
    if proc.stderr:
        print(proc.stderr.rstrip())
    return proc


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    # 2. Open workbook
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]
    cols = map_columns(ws)
    print(f"Sheet: {ws.title}")
    print(f"Columns: name={cols['name']} caption={cols['caption']} "
          f"datetime={cols['datetime']} status={cols['status']}")

    required = {k: cols[k] for k in ("name", "status")}
    missing = [k for k, v in required.items() if v is None]
    if missing:
        print(f"ERROR: missing required column(s): {missing}. "
              f"Headers: {[c.value for c in ws[1] if c.value]}")
        return

    # 3. Ensure "Video URL" column exists
    url_col = cols["video_url"]
    if url_col is None:
        # next empty header column to the right of existing headers
        used = [c.column for c in ws[1] if c.value is not None and str(c.value).strip() != ""]
        next_col = (max(used) + 1) if used else 1
        url_col = get_column_letter(next_col)
        ws[f"{url_col}1"] = "Video URL"
        wb.save(XLSX)
        print(f"Added 'Video URL' column at {url_col} and saved.")
    else:
        print(f"'Video URL' column found at {url_col}.")

    uploaded = 0
    skipped_existing = 0
    file_missing = 0

    # 4. Process each data row
    for r in range(2, ws.max_row + 1):
        name = ws[f"{cols['name']}{r}"].value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()

        raw_status = ws[f"{cols['status']}{r}"].value
        status = str(raw_status).strip().lower() if raw_status is not None else ""

        # a. skip already-handled statuses
        if status in SKIP_STATUSES:
            continue

        # b. URL already set?
        existing_url = ws[f"{url_col}{r}"].value
        if existing_url is not None and str(existing_url).strip() != "":
            print(f"row {r}: URL already set, skipping upload")
            skipped_existing += 1
            continue

        # c. upload
        video_path = os.path.join("videos", name)
        if not os.path.isfile(video_path):
            print(f"row {r}: FILE MISSING ({name}) — leaving URL blank")
            file_missing += 1
            continue

        try:
            result = cloudinary.uploader.upload_large(video_path, resource_type="video")
            secure_url = result["secure_url"]
            ws[f"{url_col}{r}"] = secure_url
            wb.save(XLSX)
            print(f"row {r}: uploaded {name} -> {secure_url}")
            uploaded += 1
        except Exception as e:
            print(f"row {r}: UPLOAD ERROR ({name}) -> {e}")
            continue

    # 5. Summary
    print("\nSummary:")
    print(f"  uploaded            : {uploaded}")
    print(f"  skipped (had URL)   : {skipped_existing}")
    print(f"  FILE MISSING        : {file_missing}")

    # 6. Git step
    print("\nGit:")
    run_git(["git", "add", "schedule.xlsx"])
    commit = run_git(["git", "commit", "-m", "Update schedule (push.py)"])
    combined = (commit.stdout + commit.stderr).lower()
    if commit.returncode != 0 and ("nothing to commit" in combined or "no changes added" in combined):
        print("No schedule changes to commit.")
    elif commit.returncode == 0:
        run_git(["git", "push"])
    else:
        print("git commit failed (see output above); skipping push.")

    # 7. Done
    print("\nPUSH COMPLETE")


if __name__ == "__main__":
    main()
