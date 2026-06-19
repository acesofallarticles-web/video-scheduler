import sys

try:
    import openpyxl
    import requests
    import cloudinary
    import dotenv
    import pytz
except ImportError as e:
    print(f"ERROR: missing package -> {e.name}. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook("schedule.xlsx")
    ws = wb["Schedule"]
except FileNotFoundError:
    print("ERROR: schedule.xlsx not found.")
    sys.exit(1)
except KeyError:
    print("ERROR: sheet 'Schedule' not found in schedule.xlsx.")
    sys.exit(1)

header = [c.value for c in ws[1]]
row = [c.value for c in ws[2]]
print("Header:", header)
print("Example row:", row)

print("SETUP OK")
