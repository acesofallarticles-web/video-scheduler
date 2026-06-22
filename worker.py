# RUNS ON GITHUB ACTIONS: posts due reels
import os
import sys
import time
from datetime import datetime

import requests
import openpyxl
import pytz
from dateutil import parser as dtparser
from dotenv import load_dotenv

GRAPH = "https://graph.facebook.com/v21.0"
IST = pytz.timezone("Asia/Kolkata")
XLSX = "schedule.xlsx"
ALREADY_STATUSES = {"posted", "published", "failed", "posting"}
POST_GAP_SECONDS = 30  # gentle pacing between consecutive posts in one run


def log(msg):
    stamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S %Z")
    print(f"[{stamp}] {msg}", flush=True)


# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
load_dotenv("config.env")
IG_ACCESS_TOKEN = os.getenv("IG_ACCESS_TOKEN")
IG_USER_ID = os.getenv("IG_USER_ID")


# ---------------------------------------------------------------------------
# WORKBOOK HELPERS
# ---------------------------------------------------------------------------
def open_ws():
    """Open workbook fresh and return (wb, ws)."""
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]
    return wb, ws


def map_columns(ws):
    headers = {}
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() != "":
            headers[cell.column_letter] = str(cell.value).strip().lower()

    def find(pred):
        for letter, text in headers.items():
            if pred(text):
                return letter
        return None

    cols = {
        "name": find(lambda h: h in ("reel name", "name")),
        "caption": find(lambda h: "caption" in h),
        "datetime": find(lambda h: "date" in h),
        "status": find(lambda h: h == "status"),
        "video_url": find(lambda h: h == "video url"),
    }
    return cols


def set_status(row, value):
    """Re-open workbook, write status to the exact row, save only if it changed."""
    wb, ws = open_ws()
    cols = map_columns(ws)
    cell = ws[f"{cols['status']}{row}"]
    current = "" if cell.value is None else str(cell.value)
    if current == str(value):
        log(f"  -> status unchanged, skipping save (row {row} already '{value}')")
        return
    cell.value = value
    wb.save(XLSX)
    log(f"  -> wrote Status='{value}' to row {row} and saved {XLSX}")


# ---------------------------------------------------------------------------
# POSTING FLOW (same proven steps as post_one.py)
# ---------------------------------------------------------------------------
def _check_token_error(data):
    """If an IG response JSON signals an expired/invalid token, log a loud alert."""
    err = data.get("error", {}) if isinstance(data, dict) else {}
    code = err.get("code")
    msg = str(err.get("message", "")).lower()
    if code == 190 or "expired" in msg or "access token" in msg:
        log("*** INSTAGRAM TOKEN EXPIRED OR INVALID — REGENERATE THE 60-DAY TOKEN ***")


def post_reel(video_url, caption):
    """Create REELS container from a pre-uploaded Cloudinary URL, poll, publish.
    Returns the published media id. Raises Exception on any failure."""
    # 1. Create media container from the pre-uploaded URL
    log(f"  [post] Using pre-uploaded video URL: {video_url}")
    log("  [post] Creating media container...")
    r = requests.post(
        f"{GRAPH}/{IG_USER_ID}/media",
        params={
            "media_type": "REELS",
            "video_url": video_url,
            "caption": caption,
            "access_token": IG_ACCESS_TOKEN,
        },
        timeout=60,
    )
    data = r.json()
    if r.status_code != 200 or "id" not in data:
        _check_token_error(data)
        raise Exception(f"Container creation failed: {data}")
    container_id = data["id"]
    log(f"  [post] Container created: {container_id}")

    # 3. Poll status_code until FINISHED
    log("  [post] Polling container status...")
    for attempt in range(1, 61):  # 60 * 5s = 5 min
        r = requests.get(
            f"{GRAPH}/{container_id}",
            params={"fields": "status_code", "access_token": IG_ACCESS_TOKEN},
            timeout=60,
        )
        status = r.json()
        status_code = status.get("status_code")
        log(f"  [post] [{attempt}/60] Status: {status_code}")
        if status_code == "FINISHED":
            break
        if status_code == "ERROR":
            _check_token_error(status)
            raise Exception(f"Container processing ERROR: {status}")
        time.sleep(5)
    else:
        raise Exception("Timed out waiting for container to finish (5 min).")

    # 4. Publish
    log("  [post] Publishing...")
    r = requests.post(
        f"{GRAPH}/{IG_USER_ID}/media_publish",
        params={"creation_id": container_id, "access_token": IG_ACCESS_TOKEN},
        timeout=60,
    )
    data = r.json()
    if r.status_code != 200 or "id" not in data:
        _check_token_error(data)
        raise Exception(f"Publish failed: {data}")
    return data["id"]


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    now = datetime.now(IST)
    log(f"Worker start. Now (Asia/Kolkata): {now.strftime('%Y-%m-%d %H:%M:%S %Z')}")

    wb, ws = open_ws()
    cols = map_columns(ws)
    log(f"Sheet: {ws.title} | columns name={cols['name']} caption={cols['caption']} "
        f"datetime={cols['datetime']} status={cols['status']}")

    missing_required = [k for k in ("name", "caption", "datetime", "status", "video_url") if cols[k] is None]
    if missing_required:
        log(f"FATAL: missing required column(s): {missing_required}. Headers: "
            f"{[c.value for c in ws[1] if c.value]}")
        sys.exit(1)  # genuine config error -> non-zero

    # Scan all rows -> classify. A row is DUE if it's pending and its time is past
    # (no matter how far past). Future rows are left scheduled. Nothing is "Missed".
    due_rows = []  # (row, scheduled_dt, name, caption, video_url)
    for r in range(2, ws.max_row + 1):
        name = ws[f"{cols['name']}{r}"].value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        caption = ws[f"{cols['caption']}{r}"].value or ""
        raw_status = ws[f"{cols['status']}{r}"].value
        status = str(raw_status).strip().lower() if raw_status is not None else ""

        if status in ALREADY_STATUSES:
            if status == "posting":
                log(f"WARNING: row {r} is stuck in 'Posting' — a previous run was "
                    f"interrupted. Check Instagram manually: if the reel posted, set "
                    f"Status to 'Posted'; if not, clear the Status to reschedule.")
            continue  # already handled

        raw_dt = ws[f"{cols['datetime']}{r}"].value
        if raw_dt is None or str(raw_dt).strip() == "":
            continue  # no date -> can't schedule
        try:
            naive = dtparser.parse(str(raw_dt))
            dt = IST.localize(naive) if naive.tzinfo is None else naive.astimezone(IST)
        except Exception:
            log(f"Row {r}: BAD_DATE ({raw_dt!r}) -> skipping")
            continue

        if dt > now:
            continue  # future -> leave scheduled

        raw_url = ws[f"{cols['video_url']}{r}"].value
        video_url = str(raw_url).strip() if raw_url is not None else ""
        due_rows.append((r, dt, name, str(caption).strip(), video_url))

    # No due reels?
    if not due_rows:
        log("No reels due. Nothing to post.")
        sys.exit(0)

    # Post ALL due reels, oldest scheduled time first.
    due_rows.sort(key=lambda x: x[1])
    log(f"{len(due_rows)} due reel(s) to post this run.")

    posted = 0
    failed = 0
    for i, (row, dt, name, caption, video_url) in enumerate(due_rows):
        log(f"--- Due reel {i + 1}/{len(due_rows)}: row {row} | {name} | "
            f"scheduled {dt.strftime('%Y-%m-%d %H:%M %Z')} ---")

        # a. Lock the row immediately.
        set_status(row, "Posting")
        log(f"Marked row {row} as Posting (lock)")

        # b. Need a pre-uploaded Cloudinary URL.
        if video_url == "":
            log(f"Row {row}: no Video URL — the video was never pushed/uploaded.")
            set_status(row, "Failed (no video URL)")
            failed += 1
        else:
            # c. Post from the pre-uploaded URL, then record result.
            try:
                media_id = post_reel(video_url, caption)
                set_status(row, "Posted")
                log(f"POSTED row {row}: {name} (Media ID: {media_id})")
                posted += 1
            except Exception as e:
                log(f"Row {row}: POST FAILED -> {e}")
                set_status(row, "Failed")
                failed += 1

        # d. Gentle pacing before the next due reel (skip after the last one).
        if i < len(due_rows) - 1:
            log(f"Sleeping {POST_GAP_SECONDS}s before next post...")
            time.sleep(POST_GAP_SECONDS)

    # 7. Summary
    log(f"Run complete. Posted: {posted} | Failed: {failed}")
    sys.exit(0)


if __name__ == "__main__":
    main()
