import os
import sys
from datetime import datetime

import openpyxl
import pytz
from dateutil import parser as dtparser
from dotenv import load_dotenv

# 1. Load config (we won't post; just keeps env consistent)
load_dotenv("config.env")

IST = pytz.timezone("Asia/Kolkata")
ALREADY_STATUSES = {"posted", "published", "missed", "failed"}

# 2. Open workbook, first sheet
try:
    wb = openpyxl.load_workbook("schedule.xlsx", data_only=True)
except FileNotFoundError:
    print("ERROR: schedule.xlsx not found.")
    sys.exit(1)

ws = wb[wb.sheetnames[0]]
print(f"Sheet: {ws.title}")

# 3. Map columns by header text (case-insensitive, trimmed)
headers = {}  # column letter -> header text
for cell in ws[1]:
    if cell.value is not None and str(cell.value).strip() != "":
        headers[cell.column_letter] = str(cell.value).strip()


def find_col(predicate):
    for letter, text in headers.items():
        if predicate(text.lower()):
            return letter
    return None


col_name = find_col(lambda h: h in ("reel name", "name"))
col_caption = find_col(lambda h: "caption" in h)
col_datetime = find_col(lambda h: "date" in h)
col_status = find_col(lambda h: h == "status")

print(f"  reel name -> {col_name}")
print(f"  caption   -> {col_caption}")
print(f"  datetime  -> {col_datetime}")
print(f"  status    -> {col_status}")

required = {"reel name": col_name, "caption": col_caption, "datetime": col_datetime}
missing = [k for k, v in required.items() if v is None]
if missing:
    print(f"\nERROR: missing required column(s): {', '.join(missing)}")
    print("Headers found:", list(headers.values()))
    sys.exit(1)

# 5. Current time in IST
now = datetime.now(pytz.utc).astimezone(IST)
print(f"\nNow (Asia/Kolkata): {now.strftime('%Y-%m-%d %H:%M:%S %Z')}\n")

# 4 + 6 + 7. Process each data row
rows = []
counts = {
    "DUE NOW": 0,
    "SCHEDULED": 0,
    "MISSED": 0,
    "ALREADY": 0,
    "BAD_DATE": 0,
    "FILE MISSING": 0,
}

for r in range(2, ws.max_row + 1):
    name = ws[f"{col_name}{r}"].value
    if name is None or str(name).strip() == "":
        continue  # skip empty rows
    name = str(name).strip()

    raw_dt = ws[f"{col_datetime}{r}"].value
    raw_status = ws[f"{col_status}{r}"].value if col_status else None
    status = str(raw_status).strip().lower() if raw_status is not None else ""

    # a. parse datetime
    parsed_str = "BAD_DATE"
    dt = None
    if raw_dt is not None and str(raw_dt).strip() != "":
        try:
            naive = dtparser.parse(str(raw_dt))
            dt = IST.localize(naive) if naive.tzinfo is None else naive.astimezone(IST)
            parsed_str = dt.strftime("%Y-%m-%d %H:%M %Z")
        except Exception:
            dt = None
            parsed_str = "BAD_DATE"

    # 6. classify
    if dt is None:
        classification = "BAD_DATE"
    elif status in ALREADY_STATUSES:
        classification = f"ALREADY: {str(raw_status).strip()}"
    else:
        delta_min = (now - dt).total_seconds() / 60.0  # positive => past
        if delta_min < 0:
            classification = "SCHEDULED"
        elif delta_min <= 10:
            classification = "DUE NOW"
        else:
            classification = "MISSED"

    # 7. file check
    file_path = os.path.join("videos", name)
    file_check = "FILE OK" if os.path.isfile(file_path) else "FILE MISSING"

    # tally
    if classification == "BAD_DATE":
        counts["BAD_DATE"] += 1
    elif classification.startswith("ALREADY"):
        counts["ALREADY"] += 1
    elif classification in counts:
        counts[classification] += 1
    if file_check == "FILE MISSING":
        counts["FILE MISSING"] += 1

    rows.append((r, name, parsed_str, classification, file_check))

# 8. aligned table
if rows:
    w_name = max(9, max(len(x[1]) for x in rows))
    w_dt = max(13, max(len(x[2]) for x in rows))
    w_cls = max(14, max(len(x[3]) for x in rows))
    header_line = f"{'Row':>4}  {'Reel Name':<{w_name}}  {'Datetime (IST)':<{w_dt}}  {'Classification':<{w_cls}}  File"
    print(header_line)
    print("-" * len(header_line))
    for r, name, parsed_str, classification, file_check in rows:
        print(f"{r:>4}  {name:<{w_name}}  {parsed_str:<{w_dt}}  {classification:<{w_cls}}  {file_check}")
else:
    print("(no data rows found)")

# 9. summary
print("\nSummary:")
print(f"  DUE NOW      : {counts['DUE NOW']}")
print(f"  SCHEDULED    : {counts['SCHEDULED']}")
print(f"  MISSED       : {counts['MISSED']}")
print(f"  ALREADY      : {counts['ALREADY']}")
print(f"  BAD_DATE     : {counts['BAD_DATE']}")
print(f"  FILE MISSING : {counts['FILE MISSING']}")
